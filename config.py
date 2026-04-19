"""
config.py — App-wide constants and configuration
"""

from openpyxl.styles import PatternFill, Font

# ── Column names ──────────────────────────────────────────────────────────────
SYMBOL_COL   = "NSE / BSE Symbol"
EXCHANGE_COL = "Exchange"

# ── Yahoo Finance ─────────────────────────────────────────────────────────────
DELAY = 0.1  # seconds between requests

SYMBOL_OVERRIDES = {
    "BAJAJ-AUTO": "BAJAJ-AUTO.NS",
    "M&M":        "M&M.NS",
    "M&MFIN":     "M&MFIN.NS",
    "TATAMTRDVR": "TATAMTRDVR.NS",
}

# ── Excel styles ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1A3C5E")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
MISS_FILL   = PatternFill("solid", fgColor="FFEB9C")
PRICE_FONT  = Font(name="Calibri", size=10)

# Alternating header colors used for date-range columns
ALT_HEADERS = [
    "2E4057", "1A6B5E", "5E3A1A", "3A1A5E", "1A5E3A",
]
