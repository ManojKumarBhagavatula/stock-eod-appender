"""
processors.py — Core processing logic for each of the four fetch modes
"""

import time
from datetime import date, datetime, timedelta

import yfinance as yf
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from config import (
    DELAY, HEADER_FONT,
    HEADER_FILL, GREEN_FILL, RED_FILL, MISS_FILL, PRICE_FONT,
    ALT_HEADERS,
)
from excel_utils import load_and_validate, wb_to_bytes
from fetcher import build_ticker, fetch_current_price, fetch_price_on_date, fetch_price_range


# ── Mode 1: Append today's EOD prices ────────────────────────────────────────
def process_eod(file_bytes: bytes) -> tuple[bytes, dict]:
    wb, ws, headers, sym_idx, exch_idx = load_and_validate(file_bytes)

    today_str   = date.today().strftime("%d-%b-%Y")
    new_col_idx = len(headers) + 1

    for cell in ws[1]:
        if str(cell.value) == today_str:
            raise ValueError(f"Today's column '{today_str}' already exists in this file.")

    h_cell = ws.cell(row=1, column=new_col_idx, value=today_str)
    h_cell.fill = HEADER_FILL
    h_cell.font = HEADER_FONT

    prev_col_idx = new_col_idx - 1
    prev_prices: dict[int, object] = {}
    if prev_col_idx > len(headers):
        for row in ws.iter_rows(min_row=2, min_col=prev_col_idx,
                                max_col=prev_col_idx, values_only=False):
            prev_prices[row[0].row] = row[0].value

    hit, miss, errors = 0, 0, []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        symbol   = row[sym_idx].value
        exchange = row[exch_idx].value
        row_num  = row[0].row
        if not symbol:
            continue

        ticker = build_ticker(str(symbol).strip(), str(exchange or "NSE").strip())
        price  = fetch_current_price(ticker)
        cell   = ws.cell(row=row_num, column=new_col_idx)
        cell.font = PRICE_FONT

        if price is None:
            cell.value = "N/A"
            cell.fill  = MISS_FILL
            miss += 1
            errors.append(symbol)
        else:
            cell.value         = price
            cell.number_format = "#,##0.00"
            hit += 1
            prev = prev_prices.get(row_num)
            if prev and isinstance(prev, (int, float)):
                cell.fill = GREEN_FILL if price >= prev else RED_FILL

        time.sleep(DELAY)

    ws.column_dimensions[get_column_letter(new_col_idx)].width = 14

    stats = {
        "date":           today_str,
        "total":          hit + miss,
        "fetched":        hit,
        "failed":         miss,
        "failed_symbols": errors[:10],
    }
    return wb_to_bytes(wb), stats


# ── Mode 2: Get current / live prices ────────────────────────────────────────
def process_current(file_bytes: bytes) -> tuple[bytes, dict]:
    wb, ws, headers, sym_idx, exch_idx = load_and_validate(file_bytes)

    now_str     = datetime.now().strftime("%d-%b-%Y %H:%M")
    col_label   = f"Live {now_str}"
    new_col_idx = len(headers) + 1

    h_cell = ws.cell(row=1, column=new_col_idx, value=col_label)
    h_cell.fill = PatternFill("solid", fgColor="0F4C81")
    h_cell.font = HEADER_FONT

    prev_col_idx = new_col_idx - 1
    prev_prices: dict[int, object] = {}
    if prev_col_idx > len(headers):
        for row in ws.iter_rows(min_row=2, min_col=prev_col_idx,
                                max_col=prev_col_idx, values_only=False):
            prev_prices[row[0].row] = row[0].value

    hit, miss, errors = 0, 0, []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        symbol   = row[sym_idx].value
        exchange = row[exch_idx].value
        row_num  = row[0].row
        if not symbol:
            continue

        ticker = build_ticker(str(symbol).strip(), str(exchange or "NSE").strip())
        price  = fetch_current_price(ticker)
        cell   = ws.cell(row=row_num, column=new_col_idx)
        cell.font = PRICE_FONT

        if price is None:
            cell.value = "N/A"
            cell.fill  = MISS_FILL
            miss += 1
            errors.append(symbol)
        else:
            cell.value         = price
            cell.number_format = "#,##0.00"
            hit += 1
            prev = prev_prices.get(row_num)
            if prev and isinstance(prev, (int, float)):
                cell.fill = GREEN_FILL if price >= prev else RED_FILL

        time.sleep(DELAY)

    ws.column_dimensions[get_column_letter(new_col_idx)].width = 14

    stats = {
        "date":           now_str,
        "total":          hit + miss,
        "fetched":        hit,
        "failed":         miss,
        "failed_symbols": errors[:10],
    }
    return wb_to_bytes(wb), stats


# ── Mode 3: Closing price on a specific date ──────────────────────────────────
def process_single_date(file_bytes: bytes, target: date) -> tuple[bytes, dict]:
    wb, ws, headers, sym_idx, exch_idx = load_and_validate(file_bytes)

    date_str    = target.strftime("%d-%b-%Y")
    new_col_idx = len(headers) + 1

    h_cell = ws.cell(row=1, column=new_col_idx, value=date_str)
    h_cell.fill = PatternFill("solid", fgColor="4B1A6B")
    h_cell.font = HEADER_FONT

    prev_col_idx = new_col_idx - 1
    prev_prices: dict[int, object] = {}
    if prev_col_idx > len(headers):
        for row in ws.iter_rows(min_row=2, min_col=prev_col_idx,
                                max_col=prev_col_idx, values_only=False):
            prev_prices[row[0].row] = row[0].value

    hit, miss, errors = 0, 0, []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        symbol   = row[sym_idx].value
        exchange = row[exch_idx].value
        row_num  = row[0].row
        if not symbol:
            continue

        ticker = build_ticker(str(symbol).strip(), str(exchange or "NSE").strip())
        price  = fetch_price_on_date(ticker, target)
        cell   = ws.cell(row=row_num, column=new_col_idx)
        cell.font = PRICE_FONT

        if price is None:
            cell.value = "N/A"
            cell.fill  = MISS_FILL
            miss += 1
            errors.append(symbol)
        else:
            cell.value         = price
            cell.number_format = "#,##0.00"
            hit += 1
            prev = prev_prices.get(row_num)
            if prev and isinstance(prev, (int, float)):
                cell.fill = GREEN_FILL if price >= prev else RED_FILL

        time.sleep(DELAY)

    ws.column_dimensions[get_column_letter(new_col_idx)].width = 14

    stats = {
        "date":           date_str,
        "total":          hit + miss,
        "fetched":        hit,
        "failed":         miss,
        "failed_symbols": errors[:10],
    }
    return wb_to_bytes(wb), stats


# ── Mode 4: Prices for a date range ──────────────────────────────────────────
def process_range(file_bytes: bytes, from_date: date, to_date: date) -> tuple[bytes, dict]:
    wb, ws, headers, sym_idx, exch_idx = load_and_validate(file_bytes)

    # Determine actual trading days from the first valid stock's history
    ref_ticker = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        sym  = row[sym_idx].value
        exch = row[exch_idx].value
        if sym:
            ref_ticker = build_ticker(str(sym).strip(), str(exch or "NSE").strip())
            break

    trading_days: list[str] = []
    if ref_ticker:
        try:
            end  = to_date + timedelta(days=1)
            hist = yf.Ticker(ref_ticker).history(
                start=from_date.isoformat(), end=end.isoformat()
            )
            trading_days = [idx.strftime("%d-%b-%Y") for idx in hist.index]
        except Exception:
            pass

    if not trading_days:
        raise ValueError("No trading days found in the selected range. Try a wider range or check dates.")

    base_col = len(headers) + 1

    for i, day_label in enumerate(trading_days):
        col_idx = base_col + i
        color   = ALT_HEADERS[i % len(ALT_HEADERS)]
        h_cell  = ws.cell(row=1, column=col_idx, value=day_label)
        h_cell.fill = PatternFill("solid", fgColor=color)
        h_cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    hit, miss, errors = 0, 0, []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        symbol   = row[sym_idx].value
        exchange = row[exch_idx].value
        row_num  = row[0].row
        if not symbol:
            continue

        ticker     = build_ticker(str(symbol).strip(), str(exchange or "NSE").strip())
        day_prices = fetch_price_range(ticker, from_date, to_date)

        prev_price = None
        for i, day_label in enumerate(trading_days):
            col_idx = base_col + i
            price   = day_prices.get(day_label)
            cell    = ws.cell(row=row_num, column=col_idx)
            cell.font = PRICE_FONT

            if price is None:
                cell.value = "N/A"
                cell.fill  = MISS_FILL
                if i == 0:
                    miss += 1
                    errors.append(symbol)
            else:
                cell.value         = price
                cell.number_format = "#,##0.00"
                if i == 0:
                    hit += 1
                if prev_price is not None:
                    cell.fill = GREEN_FILL if price >= prev_price else RED_FILL
                prev_price = price

        time.sleep(DELAY)

    stats = {
        "date":           f"{from_date.strftime('%d-%b-%Y')} → {to_date.strftime('%d-%b-%Y')}",
        "from_date":      from_date.isoformat(),
        "to_date":        to_date.isoformat(),
        "total":          hit + miss,
        "fetched":        hit,
        "failed":         miss,
        "failed_symbols": errors[:10],
        "columns":        len(trading_days),
        "trading_days":   trading_days,
    }
    return wb_to_bytes(wb), stats