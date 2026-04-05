"""
main.py — Stock EOD Price Appender (FastAPI)
Run:  uvicorn main:app --reload
"""

import io
import time
import asyncio
from datetime import date
from typing import Optional

import yfinance as yf
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, File, UploadFile, Request, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from jinja2 import Environment, FileSystemLoader

# ── App setup ─────────────────────────────────────────────────────────────────
app = FastAPI(title="Stock EOD Appender")
env = Environment(loader=FileSystemLoader("templates"))

# ── Config ────────────────────────────────────────────────────────────────────
SYMBOL_COL   = "NSE / BSE Symbol"
EXCHANGE_COL = "Exchange"
DELAY        = 0.1  # seconds between Yahoo Finance requests

SYMBOL_OVERRIDES = {
    "BAJAJ-AUTO": "BAJAJ-AUTO.NS",
    "M&M":        "M&M.NS",
    "M&MFIN":     "M&MFIN.NS",
    "TATAMTRDVR": "TATAMTRDVR.NS",
}

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1A3C5E")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
MISS_FILL   = PatternFill("solid", fgColor="FFEB9C")
PRICE_FONT  = Font(name="Calibri", size=10)

# ── Helpers ───────────────────────────────────────────────────────────────────
def build_ticker(symbol: str, exchange: str) -> str:
    if symbol in SYMBOL_OVERRIDES:
        return SYMBOL_OVERRIDES[symbol]
    suffix = ".NS" if exchange.upper() == "NSE" else ".BO"
    return symbol + suffix


def fetch_price(ticker: str) -> Optional[float]:
    try:
        info = yf.Ticker(ticker).fast_info
        price = getattr(info, "last_price", None) or getattr(info, "regular_market_price", None)
        if price:
            return round(float(price), 2)
        hist = yf.Ticker(ticker).history(period="2d")
        if not hist.empty:
            return round(float(hist["Close"].iloc[-1]), 2)
    except Exception:
        pass
    return None


def process_workbook(file_bytes: bytes) -> tuple[bytes, dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # Find headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if SYMBOL_COL not in headers or EXCHANGE_COL not in headers:
        raise ValueError(
            f"Required columns not found. Expected '{SYMBOL_COL}' and '{EXCHANGE_COL}'."
        )

    sym_idx  = headers.index(SYMBOL_COL)
    exch_idx = headers.index(EXCHANGE_COL)

    today_str   = date.today().strftime("%d-%b-%Y")
    new_col_idx = len(headers) + 1

    # Guard: don't duplicate today's column
    for cell in ws[1]:
        if str(cell.value) == today_str:
            raise ValueError(f"Today's column '{today_str}' already exists in this file.")

    # Header cell
    h_cell = ws.cell(row=1, column=new_col_idx, value=today_str)
    h_cell.fill = HEADER_FILL
    h_cell.font = HEADER_FONT

    # Previous date column for color comparison
    prev_col_idx = new_col_idx - 1
    prev_prices  = {}
    if prev_col_idx > len(headers):
        for row in ws.iter_rows(min_row=2, min_col=prev_col_idx,
                                max_col=prev_col_idx, values_only=False):
            prev_prices[row[0].row] = row[0].value

    hit, miss, errors = 0, 0, []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        symbol   = row[sym_idx].value
        exchange = row[exch_idx].value
        row_num  = row[0].row

        if not symbol:
            continue

        ticker = build_ticker(str(symbol).strip(), str(exchange or "NSE").strip())
        price  = fetch_price(ticker)
        cell   = ws.cell(row=row_num, column=new_col_idx)
        cell.font = PRICE_FONT

        if price is None:
            cell.value = "N/A"
            cell.fill  = MISS_FILL
            miss += 1
            errors.append(symbol)
        else:
            cell.value = price
            cell.number_format = "#,##0.00"
            hit += 1
            prev = prev_prices.get(row_num)
            if prev and isinstance(prev, (int, float)):
                cell.fill = GREEN_FILL if price >= prev else RED_FILL

        time.sleep(DELAY)

    ws.column_dimensions[get_column_letter(new_col_idx)].width = 14

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    stats = {
        "date": today_str,
        "total": hit + miss,
        "fetched": hit,
        "failed": miss,
        "failed_symbols": errors[:10],  # first 10 only
    }
    return out.read(), stats


# ── Routes ────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def index(request: Request): 
    template = env.get_template("index.html")
    return HTMLResponse(content=template.render())


@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 20 MB).")

    try:
        result_bytes, stats = await asyncio.to_thread(process_workbook, contents)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing error: {str(e)}")

    filename = f"stocks_eod_{stats['date'].replace('-', '_')}.xlsx"

    return StreamingResponse(
        io.BytesIO(result_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "X-Stats-Date":    stats["date"],
            "X-Stats-Fetched": str(stats["fetched"]),
            "X-Stats-Failed":  str(stats["failed"]),
            "X-Stats-Total":   str(stats["total"]),
        },
    )
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000)
